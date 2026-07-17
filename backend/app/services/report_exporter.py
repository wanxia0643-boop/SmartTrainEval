"""Generate Excel and printable PDF report files."""
from datetime import datetime
from pathlib import Path
from uuid import uuid4
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.access import is_admin, is_teacher
from app.models.achievement import TrainAchievement
from app.models.eval_result import EvalResult
from app.models.llm_log import LlmCallLog
from app.models.project import TrainProject
from app.models.user import User
from app.schemas.auth import CurrentUser

BACKEND_DIR = Path(__file__).resolve().parents[2]
REPORT_ROOT = BACKEND_DIR / "generated" / "reports"


def _report_path(prefix: str, suffix: str) -> tuple[Path, str]:
    REPORT_ROOT.mkdir(parents=True, exist_ok=True)
    filename = f"{prefix}-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid4().hex[:8]}.{suffix}"
    path = REPORT_ROOT / filename
    return path, f"/generated/reports/{filename}"


def _cell_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value


def _table_data(rows: list[dict]) -> tuple[list[str], list[list]]:
    headers = list(rows[0].keys()) if rows else ["message"]
    values = [[_cell_value(row.get(header)) for header in headers] for row in rows]
    if not values:
        values = [["暂无数据"]]
    return headers, values


def _write_excel(rows: list[dict], prefix: str) -> str:
    path, file_url = _report_path(prefix, "xlsx")
    headers, values = _table_data(rows)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "智训评报表"
    sheet.freeze_panes = "A2"
    sheet.append(headers)
    for row in values:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="0A152D")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 24
    sheet.auto_filter.ref = sheet.dimensions

    for index, header in enumerate(headers, start=1):
        content_width = max(
            [len(str(header)), *[len(str(row[index - 1])) for row in values]],
        )
        sheet.column_dimensions[get_column_letter(index)].width = min(max(content_width + 3, 12), 36)
        for cell in sheet[get_column_letter(index)]:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    workbook.save(path)
    return file_url


def _write_pdf(rows: list[dict], prefix: str) -> str:
    path, file_url = _report_path(prefix, "pdf")
    headers, values = _table_data(rows)
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ChineseTitle",
        parent=styles["Title"],
        fontName="STSong-Light",
        fontSize=16,
        leading=20,
        textColor=colors.HexColor("#0A152D"),
    )
    cell_style = ParagraphStyle(
        "ChineseCell",
        parent=styles["BodyText"],
        fontName="STSong-Light",
        fontSize=7,
        leading=9,
        wordWrap="CJK",
    )
    document = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="智训评实训评价报表",
    )
    usable_width = landscape(A4)[0] - 24 * mm
    column_widths = [usable_width / len(headers)] * len(headers)
    table_rows = [
        [Paragraph(escape(str(header)), cell_style) for header in headers],
        *[[Paragraph(escape(str(value)), cell_style) for value in row] for row in values],
    ]
    table = Table(table_rows, colWidths=column_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A152D")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), .35, colors.HexColor("#D9E1EA")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    document.build([
        Paragraph("智训评实训评价报表", title_style),
        Spacer(1, 5 * mm),
        table,
    ])
    return file_url


def _write_report(rows: list[dict], prefix: str, file_format: str) -> str:
    if file_format == "PDF":
        return _write_pdf(rows, prefix)
    return _write_excel(rows, prefix)


def _achievement_query(db: Session, current: CurrentUser, project_id: int | None):
    stmt = (
        select(TrainAchievement, TrainProject, User)
        .join(TrainProject, TrainAchievement.project_id == TrainProject.id)
        .outerjoin(User, TrainAchievement.student_id == User.id)
        .where(
            TrainAchievement.is_deleted == 0,
            TrainProject.is_deleted == 0,
        )
        .order_by(TrainAchievement.id.desc())
    )
    if project_id is not None:
        stmt = stmt.where(TrainAchievement.project_id == project_id)
    if is_teacher(current):
        stmt = stmt.where(TrainProject.teacher_id == current.user_id)
    return db.execute(stmt).all()


def generate_report_file(
    db: Session,
    *,
    report_type: int,
    project_id: int | None,
    file_format: str,
    current: CurrentUser,
) -> str:
    """Generate the requested report format and return its file URL."""
    if report_type in (1, 2, 3):
        rows = []
        for achievement, project, student in _achievement_query(db, current, project_id):
            rows.append({
                "项目": project.project_name,
                "项目编码": project.project_code,
                "学生": student.real_name if student else achievement.student_id,
                "成果标题": achievement.title,
                "状态": achievement.status,
                "最终得分": achievement.final_score if achievement.final_score is not None else "",
                "提交时间": achievement.submit_time or achievement.create_time,
            })
        return _write_report(rows, "achievement-report", file_format)

    if report_type == 4:
        stmt = select(LlmCallLog).where(LlmCallLog.is_deleted == 0).order_by(LlmCallLog.id.desc())
        if not is_admin(current):
            stmt = stmt.where(LlmCallLog.user_id == current.user_id)
        logs = db.scalars(stmt.limit(500)).all()
        rows = [{
            "业务类型": log.biz_type or "",
            "业务ID": log.biz_id or "",
            "模型": log.model_name,
            "状态": "成功" if log.status == 1 else "失败",
            "Token": log.total_tokens,
            "耗时ms": log.duration_ms,
            "错误": log.error_msg or "",
            "时间": log.create_time,
        } for log in logs]
        return _write_report(rows, "ai-usage-report", file_format)

    return _write_report([], "empty-report", file_format)
